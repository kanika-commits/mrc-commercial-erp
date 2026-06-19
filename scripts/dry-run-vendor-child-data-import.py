#!/usr/bin/env python3

import argparse
import csv
import json
import os
import re
import urllib.parse
import urllib.request
from pathlib import Path

import openpyxl


DEFAULT_WORKBOOK = "import-data/Final Work Orders - MRC _ GLC _ PI (3).xlsx"
DEFAULT_OUT_DIR = "reports/vendor-child-data-import"
SHEET_NAME = "Contractor Info"

GST_STATE_NAMES = {
    "01": "Jammu and Kashmir",
    "02": "Himachal Pradesh",
    "03": "Punjab",
    "04": "Chandigarh",
    "05": "Uttarakhand",
    "06": "Haryana",
    "07": "Delhi",
    "08": "Rajasthan",
    "09": "Uttar Pradesh",
    "10": "Bihar",
    "11": "Sikkim",
    "12": "Arunachal Pradesh",
    "13": "Nagaland",
    "14": "Manipur",
    "15": "Mizoram",
    "16": "Tripura",
    "17": "Meghalaya",
    "18": "Assam",
    "19": "West Bengal",
    "20": "Jharkhand",
    "21": "Odisha",
    "22": "Chhattisgarh",
    "23": "Madhya Pradesh",
    "24": "Gujarat",
    "27": "Maharashtra",
    "29": "Karnataka",
    "32": "Kerala",
    "33": "Tamil Nadu",
    "36": "Telangana",
    "37": "Andhra Pradesh",
}


def load_env():
    for name in [".env.local", ".env"]:
        path = Path(name)
        if not path.exists():
            continue
        for line in path.read_text(encoding="utf-8").splitlines():
            match = re.match(r"\s*([A-Z0-9_]+)\s*=\s*(.*)\s*$", line)
            if not match or match.group(1) in os.environ:
                continue
            os.environ[match.group(1)] = match.group(2).strip().strip("'\"")


def normalize_header(value):
    return re.sub(r"(^_+|_+$)", "", re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()))


def normalize_name(value):
    text = str(value or "").lower().replace("&", " and ")
    text = re.sub(r"\b(pvt|private)\b", "private", text)
    text = re.sub(r"\b(ltd|limited)\b", "limited", text)
    text = re.sub(r"[^a-z0-9]+", " ", text).strip()
    return re.sub(r"\s+", " ", text)


def normalize_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if text.endswith(".0") and re.fullmatch(r"\d+\.0", text):
        return text[:-2]
    if text.lower() in {"n/a", "na", "none", "nil", "-"}:
        return ""
    return text


def normalize_phone(value):
    text = normalize_text(value)
    digits = re.sub(r"\D+", "", text)
    return digits or ""


def normalize_email(value):
    text = normalize_text(value).lower()
    if not text or "@" not in text:
        return ""
    return text


def normalize_gstin(value):
    text = re.sub(r"[^0-9A-Za-z]+", "", normalize_text(value)).upper()
    return text if len(text) == 15 else ""


def normalize_account(value):
    return re.sub(r"\D+", "", normalize_text(value))


def normalize_ifsc(value):
    text = re.sub(r"[^0-9A-Za-z]+", "", normalize_text(value)).upper()
    match = re.search(r"[A-Z]{4}0[A-Z0-9]{6}", text)
    return match.group(0) if match else ""


def supabase_get(table, select, extra_params=None):
    base_url = os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
    service_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not base_url or not service_key:
        raise RuntimeError("Missing NEXT_PUBLIC_SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY.")

    all_rows = []
    offset = 0
    page_size = 1000
    while True:
        params = {"select": select, "limit": str(page_size), "offset": str(offset)}
        if extra_params:
            params.update(extra_params)
        url = f"{base_url.rstrip('/')}/rest/v1/{table}?{urllib.parse.urlencode(params)}"
        request = urllib.request.Request(
            url,
            headers={
                "apikey": service_key,
                "Authorization": f"Bearer {service_key}",
                "Accept": "application/json",
            },
        )
        with urllib.request.urlopen(request) as response:
            rows = json.loads(response.read().decode("utf-8"))
        all_rows.extend(rows)
        if len(rows) < page_size:
            break
        offset += page_size
    return all_rows


def parse_bank_details(value):
    text = normalize_text(value)
    if not text:
        return {}

    lines = [line.strip() for line in re.split(r"[\r\n]+", text) if line.strip()]
    joined = "\n".join(lines)
    result = {
        "account_number": "",
        "ifsc_code": "",
        "bank_name": "",
        "branch_name": "",
    }

    account_match = re.search(
        r"(?:A/C|AC|ACCOUNT)\s*(?:NO|NUMBER)?\s*[:.\-]?\s*([0-9][0-9\s\-]{5,})",
        joined,
        re.IGNORECASE,
    )
    if account_match:
        result["account_number"] = normalize_account(account_match.group(1))

    ifsc_match = re.search(r"[A-Z]{4}0[A-Z0-9]{6}", joined.upper())
    if ifsc_match:
        result["ifsc_code"] = ifsc_match.group(0)

    for line in lines:
        normalized = line.upper()
        if "BANK" in normalized and "NAME" in normalized:
            result["bank_name"] = re.sub(
                r"^\s*BANK\s*NAME\s*[:.\-]?\s*",
                "",
                line,
                flags=re.IGNORECASE,
            ).strip()
        elif normalized.startswith("BRANCH") or "BRANCH" in normalized:
            result["branch_name"] = re.sub(
                r"^\s*BRANCH\s*[:.\-]?\s*",
                "",
                line,
                flags=re.IGNORECASE,
            ).strip()

    return result


def read_contractor_rows(workbook_path):
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise RuntimeError(f"Missing workbook sheet: {SHEET_NAME}")
    sheet = workbook[SHEET_NAME]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [normalize_header(cell) for cell in rows[0]]
    output = []
    for row_number, values in enumerate(rows[1:], start=2):
        row = {headers[index]: values[index] if index < len(values) else "" for index in range(len(headers))}
        vendor_name = normalize_text(row.get("contractor_name"))
        if not vendor_name:
            continue
        output.append(
            {
                "source_row": row_number,
                "vendor_name": vendor_name,
                "contact_name": normalize_text(row.get("contact_person")),
                "contact_number": normalize_phone(row.get("contact_number")),
                "email": normalize_email(row.get("email")),
                "designation": normalize_text(row.get("designation")),
                "gstin": normalize_gstin(row.get("gstin")),
                "bank": parse_bank_details(row.get("bank_details")),
            }
        )
    return output


def write_csv(path, rows, headers):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow({header: "" if row.get(header) is None else row.get(header, "") for header in headers})


def build_outputs(source_rows, vendors, contacts, gstins, bank_accounts):
    vendors_by_name = {}
    for vendor in vendors:
        key = normalize_name(vendor.get("vendor_name"))
        vendors_by_name.setdefault(key, []).append(vendor)

    existing_contact_keys = {
        (
            row.get("vendor_id"),
            normalize_name(row.get("contact_name")),
            normalize_phone(row.get("contact_number")),
            normalize_email(row.get("email")),
        )
        for row in contacts
    }
    existing_gstin_keys = {
        (row.get("vendor_id"), normalize_gstin(row.get("gstin"))) for row in gstins
    }
    existing_bank_keys = {
        (
            row.get("vendor_id"),
            normalize_account(row.get("account_number")),
            normalize_ifsc(row.get("ifsc_code")),
        )
        for row in bank_accounts
    }

    report_rows = []
    contact_rows = []
    gstin_rows = []
    bank_rows = []
    errors = []
    seen_contact_keys = set(existing_contact_keys)
    seen_gstin_keys = set(existing_gstin_keys)
    seen_bank_keys = set(existing_bank_keys)

    for source in source_rows:
        matches = vendors_by_name.get(normalize_name(source["vendor_name"]), [])
        if not matches:
            errors.append(
                {
                    "source_row": source["source_row"],
                    "vendor_name": source["vendor_name"],
                    "section": "vendor",
                    "reason": "No matching vendor found.",
                }
            )
            continue
        if len(matches) > 1:
            errors.append(
                {
                    "source_row": source["source_row"],
                    "vendor_name": source["vendor_name"],
                    "section": "vendor",
                    "reason": "Multiple matching vendors found.",
                }
            )
            continue

        vendor = matches[0]
        organization_id = vendor.get("organization_id")
        vendor_id = vendor.get("id")

        if source["contact_name"] and (source["contact_number"] or source["email"]):
            key = (
                vendor_id,
                normalize_name(source["contact_name"]),
                source["contact_number"],
                source["email"],
            )
            status = "duplicate_contact" if key in seen_contact_keys else "ready_contact"
            if status == "ready_contact":
                seen_contact_keys.add(key)
                contact_rows.append(
                    {
                        "organization_id": organization_id,
                        "vendor_id": vendor_id,
                        "vendor_name": vendor.get("vendor_name"),
                        "contact_name": source["contact_name"],
                        "contact_number": source["contact_number"],
                        "email": source["email"],
                        "designation": source["designation"],
                        "is_primary": "true",
                        "source_row": source["source_row"],
                    }
                )
            report_rows.append(
                {
                    "source_row": source["source_row"],
                    "vendor_name": source["vendor_name"],
                    "matched_vendor_id": vendor_id,
                    "section": "contact",
                    "status": status,
                    "value": source["contact_name"],
                }
            )

        if source["gstin"]:
            key = (vendor_id, source["gstin"])
            status = "duplicate_gstin" if key in seen_gstin_keys else "ready_gstin"
            if status == "ready_gstin":
                seen_gstin_keys.add(key)
                state_code = source["gstin"][:2]
                gstin_rows.append(
                    {
                        "organization_id": organization_id,
                        "vendor_id": vendor_id,
                        "vendor_name": vendor.get("vendor_name"),
                        "gstin": source["gstin"],
                        "state_code": state_code,
                        "state_name": GST_STATE_NAMES.get(state_code, ""),
                        "is_primary": "true",
                        "source_row": source["source_row"],
                    }
                )
            report_rows.append(
                {
                    "source_row": source["source_row"],
                    "vendor_name": source["vendor_name"],
                    "matched_vendor_id": vendor_id,
                    "section": "gstin",
                    "status": status,
                    "value": source["gstin"],
                }
            )

        bank = source["bank"]
        if bank.get("account_number") or bank.get("ifsc_code") or bank.get("bank_name"):
            if not bank.get("account_number") or not bank.get("ifsc_code"):
                errors.append(
                    {
                        "source_row": source["source_row"],
                        "vendor_name": source["vendor_name"],
                        "section": "bank",
                        "reason": "Bank details are missing account number or IFSC.",
                    }
                )
                continue
            key = (vendor_id, bank["account_number"], bank["ifsc_code"])
            status = "duplicate_bank" if key in seen_bank_keys else "ready_bank"
            if status == "ready_bank":
                seen_bank_keys.add(key)
                bank_rows.append(
                    {
                        "organization_id": organization_id,
                        "vendor_id": vendor_id,
                        "vendor_name": vendor.get("vendor_name"),
                        "account_holder_name": vendor.get("vendor_name"),
                        "account_number": bank["account_number"],
                        "ifsc_code": bank["ifsc_code"],
                        "bank_name": bank.get("bank_name") or "",
                        "branch_name": bank.get("branch_name") or "",
                        "is_primary": "true",
                        "source_row": source["source_row"],
                    }
                )
            report_rows.append(
                {
                    "source_row": source["source_row"],
                    "vendor_name": source["vendor_name"],
                    "matched_vendor_id": vendor_id,
                    "section": "bank",
                    "status": status,
                    "value": bank.get("account_number") or bank.get("bank_name") or "",
                }
            )

    return report_rows, contact_rows, gstin_rows, bank_rows, errors


def parse_args():
    parser = argparse.ArgumentParser(description="Dry-run vendor child data import from Contractor Info.")
    parser.add_argument("--source", default=DEFAULT_WORKBOOK)
    parser.add_argument("--out-dir", default=DEFAULT_OUT_DIR)
    return parser.parse_args()


def main():
    args = parse_args()
    load_env()

    workbook_path = Path(args.source).expanduser().resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    source_rows = read_contractor_rows(workbook_path)
    vendors = supabase_get("vendors", "id,organization_id,vendor_name,gstin")
    contacts = supabase_get("vendor_contacts", "id,vendor_id,contact_name,contact_number,email")
    gstins = supabase_get("vendor_gstins", "id,vendor_id,gstin")
    bank_accounts = supabase_get("vendor_bank_accounts", "id,vendor_id,account_number,ifsc_code")

    report_rows, contact_rows, gstin_rows, bank_rows, errors = build_outputs(
        source_rows, vendors, contacts, gstins, bank_accounts
    )

    out_dir = Path(args.out_dir).resolve()
    write_csv(
        out_dir / "vendor_child_data_dry_run_report.csv",
        report_rows,
        ["source_row", "vendor_name", "matched_vendor_id", "section", "status", "value"],
    )
    write_csv(
        out_dir / "vendor_contacts_sql_ready.csv",
        contact_rows,
        [
            "organization_id",
            "vendor_id",
            "vendor_name",
            "contact_name",
            "contact_number",
            "email",
            "designation",
            "is_primary",
            "source_row",
        ],
    )
    write_csv(
        out_dir / "vendor_gstins_sql_ready.csv",
        gstin_rows,
        [
            "organization_id",
            "vendor_id",
            "vendor_name",
            "gstin",
            "state_code",
            "state_name",
            "is_primary",
            "source_row",
        ],
    )
    write_csv(
        out_dir / "vendor_bank_accounts_sql_ready.csv",
        bank_rows,
        [
            "organization_id",
            "vendor_id",
            "vendor_name",
            "account_holder_name",
            "account_number",
            "ifsc_code",
            "bank_name",
            "branch_name",
            "is_primary",
            "source_row",
        ],
    )
    write_csv(
        out_dir / "vendor_child_data_errors.csv",
        errors,
        ["source_row", "vendor_name", "section", "reason"],
    )

    print("vendor child data dry-run complete")
    print(
        json.dumps(
            {
                "source_rows": len(source_rows),
                "ready_contacts": len(contact_rows),
                "ready_gstins": len(gstin_rows),
                "ready_bank_accounts": len(bank_rows),
                "errors": len(errors),
                "out_dir": str(out_dir),
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
